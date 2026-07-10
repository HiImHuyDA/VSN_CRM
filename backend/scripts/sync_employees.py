"""
sync_employees.py
-----------------
Script Python đọc file Excel nhân viên từ ổ mạng 10.0.0.214,
so sánh ngày modified của file, nếu thay đổi thì upsert vào SQL Server.

Chạy bởi:
  - Node.js backend (khi người dùng bấm nút "Đồng bộ")
  - Có thể schedule Task Scheduler chạy tự động mỗi ngày
  
Kết quả trả về JSON stdout để Node.js đọc.
"""

import sys
import os
import json
import shutil
import tempfile
import time
from datetime import datetime
from pathlib import Path

# Thêm thư viện cần: pip install openpyxl pyodbc python-dotenv
try:
    import openpyxl
    import pyodbc
    from dotenv import load_dotenv
except ImportError as e:
    print(json.dumps({"success": False, "error": f"Thiếu thư viện: {e}. Chạy: pip install openpyxl pyodbc python-dotenv"}))
    sys.exit(1)

# ── Config ────────────────────────────────────────────────────────────────────
# Load .env từ thư mục backend
ENV_PATH = Path(__file__).parent.parent / ".env"
load_dotenv(ENV_PATH)

EXCEL_PATH = Path(r"//10.0.0.214/public/07- DANH BẠ ĐIỆN THOẠI, EMAIL…/1. VSN - DANH BẠ ĐIỆN THOẠI.xlsx")
SHEET_NAME = "DANH BẠ"
HEADER_ROW = 3  # Header ở dòng 3

DB_SERVER   = os.getenv("DB_SERVER", "10.0.0.36")
DB_PORT     = os.getenv("DB_PORT", "1433")
DB_DATABASE = os.getenv("DB_DATABASE", "CSR_DB")
DB_USER     = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")


def ensure_network_share(share_path: str):
    """Đảm bảo đường dẫn ổ mạng được xác thực trên Windows."""
    if not share_path.startswith(r"\\"):
        return
    # Trích xuất server share (ví dụ: \\10.0.0.214\public)
    parts = share_path.split("\\")
    if len(parts) >= 4:
        server_share = "\\\\" + parts[2] + "\\" + parts[3]
    else:
        server_share = "\\\\" + parts[2]
        
    # Thử check xem truy cập được chưa
    if os.path.exists(share_path):
        return

    # Nếu không truy cập được, thử chạy 'net use' để Windows nạp thông tin xác thực
    # Hỗ trợ lấy user/pass từ file .env nếu có cấu hình
    share_user = os.getenv("NET_SHARE_USER")
    share_pass = os.getenv("NET_SHARE_PASSWORD")
    
    import subprocess
    try:
        if share_user and share_pass:
            cmd = f'net use "{server_share}" "{share_pass}" /user:"{share_user}" /persistent:no'
        else:
            # Thử kết nối bằng quyền của tài khoản hiện tại
            cmd = f'net use "{server_share}" /persistent:no'
        
        subprocess.run(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=5)
    except Exception:
        pass


def get_file_modified_at(path: Path) -> datetime | None:
    """Lấy ngày modified của file Excel."""
    try:
        ensure_network_share(str(path))
        ts = os.path.getmtime(path)
        return datetime.fromtimestamp(ts)
    except Exception:
        return None


def read_excel(source_path: Path) -> list[dict] | None:
    """
    Copy file về temp rồi đọc (tránh lỗi mạng chập chờn).
    Logic giữ nguyên từ employee_handler.py cũ.
    """
    temp_file = None
    try:
        ensure_network_share(str(source_path))
        if not source_path.exists():
            raise FileNotFoundError(f"Không tìm thấy file: {source_path}")

        # Copy về temp
        fd, temp_file = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        shutil.copy2(source_path, temp_file)

        wb = openpyxl.load_workbook(temp_file, read_only=True, data_only=True)

        if SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"Không tìm thấy sheet '{SHEET_NAME}'")

        ws = wb[SHEET_NAME]

        # Đọc header tại dòng HEADER_ROW
        headers = next(
            ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True),
            None
        )
        if not headers:
            raise ValueError("Không đọc được header")

        # Tìm index cột
        col = {}
        for idx, name in enumerate(headers):
            if name in ("Họ", "Tên", "Mail", "MNV", "Re-Team"):
                col[name] = idx

        missing = [c for c in ("Họ", "Tên", "Mail", "MNV") if c not in col]
        if missing:
            raise ValueError(f"Thiếu cột: {missing}")

        # Đọc dữ liệu
        employees = []
        seen_emails = set()

        for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
            email = str(row[col["Mail"]] or "").strip().lower()
            mnv   = str(row[col["MNV"]]  or "").strip()

            if not email or "@" not in email or email == "0":
                continue
            if not mnv:
                continue
            if email in seen_emails:
                continue

            ho  = str(row[col["Họ"]]  or "").strip()
            ten = str(row[col["Tên"]] or "").strip()
            full_name = f"{ho} {ten}".strip()

            if not full_name:
                continue

            department = str(row[col["Re-Team"]] or "").strip() if "Re-Team" in col else ""

            seen_emails.add(email)
            employees.append({
                "fullName": full_name,
                "email":    email,
                "mnv":      mnv,
                "department": department
            })

        wb.close()
        return employees

    finally:
        if temp_file and os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except Exception:
                pass


def get_db_connection():
    """Kết nối SQL Server qua pyodbc."""
    conn_str = (
        f"DRIVER={{ODBC Driver 17 for SQL Server}};"
        f"SERVER={DB_SERVER},{DB_PORT};"
        f"DATABASE={DB_DATABASE};"
        f"UID={DB_USER};"
        f"PWD={DB_PASSWORD};"
        f"TrustServerCertificate=yes;"
    )
    return pyodbc.connect(conn_str, timeout=10)


def sync_to_sql(employees: list[dict], file_modified_at: datetime) -> dict:
    """Gọi usp_SyncEmployees để upsert nhân viên vào SQL."""
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        employees_json = json.dumps(employees, ensure_ascii=False)
        file_mod_str = file_modified_at.strftime("%Y-%m-%d %H:%M:%S")

        cursor.execute(
            "EXEC [dbo].[usp_SyncEmployees] @EmployeesJson=?, @FileModifiedAt=?",
            employees_json,
            file_mod_str
        )
        row = cursor.fetchone()
        conn.commit()

        return {
            "success":       True,
            "rowsAffected":  row[0] if row else len(employees),
            "fileModifiedAt": file_mod_str,
            "message":       f"Đồng bộ thành công {len(employees)} nhân viên",
            "missingEmployees": json.loads(row[3]) if row and len(row) > 3 and row[3] else [],
            "newEmployees": json.loads(row[4]) if row and len(row) > 4 and row[4] else []
        }
    finally:
        conn.close()


def main():
    force = "--force" in sys.argv  # Buộc sync dù file không thay đổi

    # 1. Kiểm tra file tồn tại
    file_modified_at = get_file_modified_at(EXCEL_PATH)
    if file_modified_at is None:
        print(json.dumps({
            "success": False,
            "error": f"Không truy cập được file: {EXCEL_PATH}"
        }))
        sys.exit(1)

    # 2. Đọc Excel (thử 3 lần)
    employees = None
    last_error = ""
    for attempt in range(3):
        try:
            employees = read_excel(EXCEL_PATH)
            break
        except Exception as e:
            last_error = str(e)
            if attempt < 2:
                time.sleep(1)

    if employees is None:
        print(json.dumps({"success": False, "error": last_error}))
        sys.exit(1)

    # 3. Đồng bộ vào SQL
    try:
        result = sync_to_sql(employees, file_modified_at)
        print(json.dumps(result))
    except Exception as e:
        print(json.dumps({"success": False, "error": str(e)}))
        sys.exit(1)


if __name__ == "__main__":
    main()
